"""
Scrape gravity-engine ranking boards (daily only), with optional login.

Usage:
  # First time: log in and save the session (opens a real browser window).
  python scrape_rank.py --login

  # Subsequent runs: re-uses the saved session if rank_auth.json exists.
  python scrape_rank.py               # default top 100 if logged in, 20 if not
  python scrape_rank.py --top 50      # cap at top 50 per board
  python scrape_rank.py --anon        # force anonymous run

Outputs (timestamped) under D:/claude/rank_output/:
  rank_<YYYYmmdd_HHMMSS>.json
  rank_<YYYYmmdd_HHMMSS>.csv
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path


def _app_dir() -> Path:
    """Directory where this app is running from — whether dev or frozen."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).parent


def _setup_playwright_env():
    """When packaged with PyInstaller, point Playwright at the browsers
    bundled next to the exe so it doesn't try ~/AppData/Local/ms-playwright."""
    if getattr(sys, "frozen", False):
        bundled = _app_dir() / "ms-playwright"
        if bundled.exists():
            os.environ.setdefault("PLAYWRIGHT_BROWSERS_PATH", str(bundled))


_setup_playwright_env()

from openpyxl import Workbook  # noqa: E402
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout  # noqa: E402

URL = "https://rank.gravity-engine.com/#/"
DEFAULT_OUT_DIR = Path("D:/claude/rank_output")
DEFAULT_AUTH_FILE = Path("D:/claude/rank_auth.json")

BOARD_LABELS = {
    "wx": ["人气榜", "畅销榜", "畅玩榜"],
    "douyin": ["热门榜", "畅销榜", "新游榜"],
}
PLATFORM_LABEL = {"wx": "微信小游戏", "douyin": "抖音小游戏"}
# Short prefix used in Excel sheet names (Excel caps sheet names at 31 chars).
PLATFORM_SHORT = {"wx": "微小", "douyin": "抖小"}

UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)


# ---------- parsing ----------

def parse_row(text: str, index_in_board: int) -> dict:
    """Parse one rank row's innerText blob into structured fields."""
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    rank: int | None = None
    name = category = subcategory = publisher = change = slogan = ""
    cat_rank: int | None = None

    if lines and lines[0].startswith("NO."):
        m = re.match(r"NO\.(\d+)", lines[0])
        if m:
            rank = int(m.group(1))
        lines = lines[1:]
    if rank is None:
        rank = index_in_board + 1

    if lines:
        name = lines[0]
    if len(lines) >= 2:
        m = re.match(r"^(.+?):(\d+)名(?:\s+(.+))?$", lines[1])
        if m:
            category = m.group(1)
            cat_rank = int(m.group(2))
            subcategory = (m.group(3) or "").strip()
        else:
            slogan = lines[1].strip('"“”')
    if len(lines) >= 3:
        publisher = lines[2] if lines[2] != "--" else ""
    if len(lines) >= 4:
        change = lines[3]

    return {
        "rank": rank,
        "name": name,
        "category": category,
        "category_rank": cat_rank,
        "subcategory": subcategory,
        "slogan": slogan,
        "publisher": publisher,
        "change": change,
    }


# ---------- page driving ----------

def kill_overlays(page, log=print):
    """Site pops a login-prompt dialog ~3s after load (anonymous only).
    Close any known close button, then hide remaining overlays via CSS."""
    for sel in (
        ".el-overlay .el-dialog__headerbtn",
        ".el-overlay .el-dialog__close",
        ".el-message-box__btns .el-button",
    ):
        try:
            page.locator(sel).first.click(timeout=1000)
            page.wait_for_timeout(150)
        except Exception:
            pass
    try:
        page.keyboard.press("Escape")
    except Exception:
        pass
    try:
        page.add_style_tag(content=(
            ".el-overlay,.el-overlay-dialog,.el-message-box__wrapper"
            "{display:none !important;pointer-events:none !important;}"
        ))
    except Exception:
        pass


def _scroll_board_to_load(page, board_handle, target: int,
                          max_idle_rounds: int = 4,
                          max_rounds: int = 40) -> int:
    """Scroll inside one board card to trigger lazy-load until we have
    `target` rank rows, or rows stop growing for `max_idle_rounds` ticks."""
    rows_locator = board_handle.locator(".rank-child-item")
    prev = rows_locator.count()
    idle = 0
    for _ in range(max_rounds):
        if prev >= target:
            return prev
        # Scroll the last visible row into view inside its own scroll
        # container — this works regardless of which ancestor is the
        # actual scrollable element.
        try:
            last = rows_locator.nth(prev - 1) if prev else board_handle
            last.scroll_into_view_if_needed(timeout=1500)
        except Exception:
            pass
        # Also nudge the board container itself in case it's the scroller.
        try:
            board_handle.evaluate(
                "el => { const s = el.querySelector('[class*=\"scroll\"]') "
                "|| el; s.scrollTop = s.scrollHeight; }"
            )
        except Exception:
            pass
        page.wait_for_timeout(600)
        cur = rows_locator.count()
        if cur == prev:
            idle += 1
            if idle >= max_idle_rounds:
                break
        else:
            idle = 0
        prev = cur
    return prev


def scrape_current_platform(page, platform: str, top_n: int, log=print) -> dict:
    boards_out = []
    labels = BOARD_LABELS[platform]
    board_handles = page.locator(".rank-list > div").all()
    if len(board_handles) < len(labels):
        log(f"[warn] {platform}: expected {len(labels)} boards, "
            f"found {len(board_handles)}")
    for i, label in enumerate(labels):
        if i >= len(board_handles):
            boards_out.append({"label": label, "rows": []})
            continue
        bh = board_handles[i]
        if top_n > 20:
            loaded = _scroll_board_to_load(page, bh, top_n)
        else:
            loaded = bh.locator(".rank-child-item").count()
        rows_text = bh.locator(".rank-child-item").all_inner_texts()
        # The arrow (up/down/flat) is a SVG inside `.rank-right-item`,
        # so it's invisible to innerText. Read it separately and zip
        # the result back onto each row.
        #
        # Anchored against real samples:
        #   羊了个羊:星球 上升一名 → 'up'   (站点显示绿色 ▼: 数字小=好)
        #   赵云与阿斗   下降一名 → 'down' (站点显示红色 ▲: 数字大=坏)
        #
        # We detect the SVG path by a stable substring rather than full
        # path equality — the site sometimes formats the d-string with
        # extra spaces or slightly different decimals.
        directions = bh.evaluate(r"""el => {
          const items = el.querySelectorAll('.rank-child-item');
          const out = [];
          for (const item of items) {
            const right = item.querySelector('.rank-right-item');
            if (!right) { out.push('unknown'); continue; }
            const text = (right.innerText || '').trim();
            if (text.indexOf('霸榜') >= 0) { out.push('top'); continue; }
            if (text.indexOf('稳定') >= 0) { out.push('flat'); continue; }
            if (text === '--' || text === '—') { out.push('new'); continue; }
            const path = right.querySelector('path');
            if (path) {
              const d = (path.getAttribute('d') || '').replace(/\s+/g, ' ').trim();
              // Up triangle (▲): "M512 320 192 704h639.936z" — has '704'
              //   = rank number got bigger = worsened → 'down'
              if (d.indexOf('704') >= 0) { out.push('down'); continue; }
              // Down triangle (▼): "m192 384 320 384 320-384z" — has '-384'
              //   = rank number got smaller = improved → 'up'
              if (d.indexOf('-384') >= 0) { out.push('up'); continue; }
            }
            // Fallback: try computed color (green=up, red=down).
            const icon = right.querySelector('i.el-icon, svg');
            if (icon) {
              const c = window.getComputedStyle(icon).color || '';
              const m = c.match(/(\d+)\s*,\s*(\d+)\s*,\s*(\d+)/);
              if (m) {
                const r = +m[1], g = +m[2], b = +m[3];
                if (g > 100 && r < 100) { out.push('up'); continue; }
                if (r > 180 && g < 120) { out.push('down'); continue; }
              }
            }
            out.push(text ? 'flat' : 'unknown');
          }
          return out;
        }""")
        rows = []
        for j, t in enumerate(rows_text):
            r = parse_row(t, j)
            r["change_direction"] = (
                directions[j] if j < len(directions) else "unknown"
            )
            rows.append(r)
        rows = rows[:top_n]
        boards_out.append({"label": label, "rows": rows})
        log(f"  {PLATFORM_LABEL[platform]} / {label}: {len(rows)} rows "
            f"(loaded {loaded}, target {top_n})")
    return {"label": PLATFORM_LABEL[platform], "boards": boards_out}


def switch_to_douyin(page, log=print) -> bool:
    tab = page.locator("img[src*='douyin_tab_rank']").first
    try:
        tab.scroll_into_view_if_needed(timeout=2000)
    except Exception:
        pass
    try:
        clickable = tab.locator(
            "xpath=ancestor::*[self::div or self::a or self::button][1]"
        )
        try:
            with page.expect_response(
                lambda r: "rank/public_list" in r.url or "rank/list" in r.url,
                timeout=8000,
            ):
                clickable.click(timeout=4000)
        except PWTimeout:
            clickable.click(timeout=4000)
        page.wait_for_timeout(1200)
        return True
    except Exception as e:
        log(f"[warn] could not click 抖音 tab: {e}")
        return False


# ---------- modes ----------

def do_login(auth_file: Path = DEFAULT_AUTH_FILE,
             wait_signal=None, log=print) -> bool:
    """Open a real browser, let the user log in, then persist cookies.

    Returns True if a session was saved (the wait_signal was set while
    the page was still alive). If the user just closes the browser
    without confirming, returns False without writing the auth file.
    """
    auth_file = Path(auth_file)
    auth_file.parent.mkdir(parents=True, exist_ok=True)
    log("正在打开浏览器，请在浏览器里完成手机号 + 验证码登录…")
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False)
            ctx = browser.new_context(viewport={"width": 1400, "height": 900},
                                      user_agent=UA)
            page = ctx.new_page()
            page.goto(URL, wait_until="domcontentloaded")

            # We need TWO ways to break out of the wait: the user clicks
            # "我已登录完成" in the GUI (wait_signal), or the user just
            # closes the browser. Use threading.Event under the hood and
            # bridge the page-close event into it.
            import threading
            done = threading.Event()
            closed = {"flag": False}

            def on_page_close(_):
                closed["flag"] = True
                done.set()

            def on_ctx_close(_):
                closed["flag"] = True
                done.set()

            page.on("close", on_page_close)
            ctx.on("close", on_ctx_close)
            browser.on("disconnected", lambda _: (closed.update(flag=True),
                                                  done.set()))

            if wait_signal is not None:
                # Wait either for the GUI signal (done set externally via
                # a wrapper) or browser close. We can't compose two waits
                # cleanly, so spawn a small thread that sets `done` once
                # the external signal fires.
                def bridge():
                    try:
                        wait_signal()
                    except Exception:
                        pass
                    done.set()
                threading.Thread(target=bridge, daemon=True).start()
                done.wait()
            else:
                try:
                    input("\n>>> 登录完成后回到这里按 Enter 保存… ")
                except EOFError:
                    pass

            if closed["flag"]:
                log("浏览器已关闭，未保存登录态。")
                return False

            ctx.storage_state(path=str(auth_file))
            try:
                browser.close()
            except Exception:
                pass
        log(f"登录态已保存：{auth_file}")
        return True
    except Exception as e:
        log(f"[错误] 登录失败：{e}")
        return False


def _pick_historical_date(page, date_str: str, xhr_counter: dict, log=print):
    """Open the top-level date picker and select `date_str`.

    Element UI el-date-picker layout:
      <input placeholder="选择日期"> ← focus/click opens the panel
      <div class="el-picker-panel">
        <table class="el-date-table">
          <td class="available"><span><em>D</em></span></td>  ← click a day
        </table>
      </div>
    """
    from datetime import datetime
    baseline = xhr_counter["n"]
    log(f"[hist] 打开日期选择器，切到 {date_str}")

    # Click the placeholder input to open the panel.
    picker_input = page.locator("input[placeholder='选择日期']").first
    picker_input.click(timeout=5000)
    page.wait_for_timeout(400)

    # Element UI's date panel exposes the current view's month in a
    # header link like `<span class="el-date-picker__header-label">2026 年 7 月</span>`.
    # We navigate month-by-month if needed. To keep this simple, use the
    # arrow buttons.
    target = datetime.strptime(date_str, "%Y-%m-%d")
    for _attempt in range(24):
        header = page.locator(".el-date-picker__header-label").all()
        if not header:
            break
        header_text = " | ".join([h.inner_text() for h in header])
        # Parse "2026 年 7 月" → year 2026, month 7.
        import re
        m = re.search(r"(\d{4})\s*年\s*(\d{1,2})\s*月", header_text)
        if not m:
            break
        cur_y, cur_m = int(m.group(1)), int(m.group(2))
        if cur_y == target.year and cur_m == target.month:
            break
        if (cur_y, cur_m) < (target.year, target.month):
            page.locator(".el-picker-panel__icon-btn.arrow-right, "
                         ".el-icon-arrow-right").first.click(timeout=2000)
        else:
            page.locator(".el-picker-panel__icon-btn.arrow-left, "
                         ".el-icon-arrow-left").first.click(timeout=2000)
        page.wait_for_timeout(200)

    # Now click the target day cell. Element UI puts each day in a
    # <td class="available"> with an inner <em>N</em>. Match on the
    # exact day number, avoiding "next/prev month" cells that carry
    # class 'next-month' or 'prev-month'.
    day = target.day
    day_cell = page.locator(
        f".el-date-table td.available:not(.next-month):not(.prev-month) "
        f":text-is('{day}')"
    ).first
    try:
        day_cell.click(timeout=5000)
    except Exception:
        # Fallback: broader match on <span> / <em> containing the day.
        page.locator(f".el-date-table td:not(.next-month):not(.prev-month) "
                     f"span:has-text('{day}')").first.click(timeout=5000)

    # Wait for a new burst of rank XHRs (one per board = 3 for wx).
    deadline_ms = 12000
    elapsed = 0
    while xhr_counter["n"] < baseline + 3 and elapsed < deadline_ms:
        page.wait_for_timeout(500)
        elapsed += 500
    log(f"[hist] 切换后 rank XHRs +{xhr_counter['n'] - baseline}")
    page.wait_for_timeout(1500)


def do_scrape(top_n: int | None,
              force_anon: bool = False,
              out_dir: Path = DEFAULT_OUT_DIR,
              auth_file: Path = DEFAULT_AUTH_FILE,
              log=print,
              historical_date: str | None = None) -> dict:
    """Scrape today's ranking, or a historical date if `historical_date`
    (YYYY-MM-DD) is provided. Historical scraping requires a logged-in
    session; anonymous mode only exposes today.
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    auth_file = Path(auth_file)
    use_auth = auth_file.exists() and not force_anon
    if top_n is None:
        top_n = 100 if use_auth else 20
    log(f"模式：{'已登录' if use_auth else '匿名'} | Top {top_n}"
        + (f" | 历史日期：{historical_date}" if historical_date else ""))

    result = {
        "scraped_at": datetime.now().isoformat(timespec="seconds"),
        "period": "日榜",
        "logged_in": use_auth,
        "top_n_target": top_n,
        "source": URL,
        "platforms": {},
    }
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        kwargs = {"viewport": {"width": 1600, "height": 1200}, "user_agent": UA}
        if use_auth:
            kwargs["storage_state"] = str(auth_file)
        ctx = browser.new_context(**kwargs)
        page = ctx.new_page()

        # Track how many rank/public_list XHRs have fired. The Vue app
        # renders the initial (stale) DOM immediately from cache, then
        # fires an XHR per board and re-renders. If we read the DOM
        # before those XHRs return, we get yesterday's data.
        rank_xhr_count = {"n": 0}

        def _on_response(resp):
            url = resp.url
            if ("rank/public_list" in url or "rank/list" in url) \
                    and resp.status == 200:
                rank_xhr_count["n"] += 1

        page.on("response", _on_response)

        page.goto(URL, wait_until="domcontentloaded", timeout=60000)
        try:
            page.wait_for_selector(".rank-child-item", timeout=30000)
        except PWTimeout:
            log("[warn] 榜单条目 30s 内未出现")

        # Wait for at least 3 rank XHRs (one per WeChat board) to
        # complete before we trust the DOM. Each response triggers a
        # re-render; give Vue a beat after each.
        deadline_ms = 15000
        elapsed = 0
        while rank_xhr_count["n"] < 3 and elapsed < deadline_ms:
            page.wait_for_timeout(500)
            elapsed += 500
        log(f"[wx] rank XHRs observed before scrape: {rank_xhr_count['n']}")
        # Extra settle time for the DOM to reflect the last XHR.
        page.wait_for_timeout(1500)

        kill_overlays(page, log=log)

        # If a historical date was requested, drive the date picker to
        # jump to that day. Wait for a fresh burst of rank XHRs before
        # trusting the DOM again.
        if historical_date:
            _pick_historical_date(page, historical_date, rank_xhr_count, log)

        result["platforms"]["wx"] = scrape_current_platform(
            page, "wx", top_n, log=log
        )

        if switch_to_douyin(page, log=log):
            kill_overlays(page, log=log)
            result["platforms"]["douyin"] = scrape_current_platform(
                page, "douyin", top_n, log=log
            )
        else:
            result["platforms"]["douyin"] = {
                "label": PLATFORM_LABEL["douyin"],
                "boards": [],
                "error": "tab switch failed",
            }

        browser.close()
    return result


def write_outputs(data: dict, out_dir: Path = DEFAULT_OUT_DIR) -> tuple[Path, Path]:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = out_dir / f"rank_{stamp}.json"
    xlsx_path = out_dir / f"rank_{stamp}.xlsx"

    json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2),
                         encoding="utf-8")

    # Excel: one sheet per (platform, board), e.g. "微小人气榜", "抖小新游榜".
    wb = Workbook()
    wb.remove(wb.active)
    headers = [
        "排名", "游戏名", "分类", "分类内排名", "子类", "标语",
        "发行商", "变化",
    ]
    for plat_key, plat in data.get("platforms", {}).items():
        short = PLATFORM_SHORT.get(plat_key, plat_key)
        for board in plat.get("boards", []):
            sheet_name = f"{short}{board['label']}"[:31]
            ws = wb.create_sheet(title=sheet_name)
            ws.append(headers)
            for r in board.get("rows", []):
                ws.append([
                    r["rank"], r["name"], r["category"],
                    r["category_rank"] or "", r["subcategory"], r["slogan"],
                    r["publisher"], r["change"],
                ])
            # Reasonable column widths.
            widths = [6, 22, 10, 10, 14, 30, 28, 12]
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
            ws.freeze_panes = "A2"
    if not wb.sheetnames:
        wb.create_sheet(title="empty")
    wb.save(xlsx_path)
    return json_path, xlsx_path


# ---------- diff between two snapshots ----------

_TIMESTAMP_RE = re.compile(r"rank_(\d{8})_(\d{6})\.xlsx$", re.IGNORECASE)


def _parse_xlsx_stamp(path: Path) -> datetime | None:
    m = _TIMESTAMP_RE.search(path.name)
    if not m:
        return None
    try:
        return datetime.strptime(m.group(1) + m.group(2), "%Y%m%d%H%M%S")
    except ValueError:
        return None


def _read_xlsx_to_boards(path: Path) -> dict[str, list[dict]]:
    """Read a rank_*.xlsx file back into {sheet_name: [row_dict, ...]}."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    out: dict[str, list[dict]] = {}
    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            out[ws.title] = []
            continue
        keys = [str(h) if h is not None else "" for h in header]
        rows = []
        for r in rows_iter:
            d = {keys[i]: (r[i] if i < len(r) else None) for i in range(len(keys))}
            if d.get("游戏名"):
                rows.append(d)
        out[ws.title] = rows
    wb.close()
    return out


def _human_delta(d: datetime, e: datetime) -> str:
    secs = abs(int((e - d).total_seconds()))
    days, secs = divmod(secs, 86400)
    hours, secs = divmod(secs, 3600)
    mins = secs // 60
    parts = []
    if days:
        parts.append(f"{days}天")
    if hours:
        parts.append(f"{hours}小时")
    if mins and not days:
        parts.append(f"{mins}分钟")
    return "".join(parts) or "<1分钟"


def diff_latest_xlsx(out_dir: Path, log=print) -> Path | None:
    """Pick the two most recent rank_*.xlsx files in `out_dir` and write
    a diff workbook listing entries (game / publisher) that are new in
    the later one. Returns the diff file path, or None on failure."""
    out_dir = Path(out_dir)
    files = sorted(
        (p for p in out_dir.glob("rank_*.xlsx")
         if _parse_xlsx_stamp(p) is not None
         and not p.name.startswith("diff_")),
        key=lambda p: _parse_xlsx_stamp(p),
    )
    if len(files) < 2:
        log(f"[diff] 至少需要 2 个 rank_*.xlsx 才能对比，"
            f"当前找到 {len(files)} 个。")
        return None
    older, newer = files[-2], files[-1]
    t_old = _parse_xlsx_stamp(older)
    t_new = _parse_xlsx_stamp(newer)
    log(f"[diff] 旧：{older.name}  ({t_old})")
    log(f"[diff] 新：{newer.name}  ({t_new})")

    old_boards = _read_xlsx_to_boards(older)
    new_boards = _read_xlsx_to_boards(newer)

    wb = Workbook()
    wb.remove(wb.active)

    # Summary sheet first.
    ws_sum = wb.create_sheet(title="总览")
    ws_sum.append(["对比基准", older.name])
    ws_sum.append(["对比目标", newer.name])
    ws_sum.append(["时间间隔", _human_delta(t_old, t_new)])
    ws_sum.append([
        "时间区间",
        f"{t_old.strftime('%Y-%m-%d %H:%M')} → "
        f"{t_new.strftime('%Y-%m-%d %H:%M')}",
    ])
    ws_sum.append([])
    ws_sum.append(["榜单", "新游戏数", "新发行商数"])
    summary_start = ws_sum.max_row + 1

    headers = ["排名", "游戏名", "分类", "子类", "标语", "发行商", "变化"]
    for sheet_name in new_boards:
        new_rows = new_boards.get(sheet_name, [])
        old_rows = old_boards.get(sheet_name, [])
        old_games = {r["游戏名"] for r in old_rows if r.get("游戏名")}
        old_publishers = {r["发行商"] for r in old_rows if r.get("发行商")}

        new_games = [r for r in new_rows if r.get("游戏名") not in old_games]
        new_pub_rows = [
            r for r in new_rows
            if r.get("发行商") and r["发行商"] not in old_publishers
        ]
        # Dedup by publisher in new_pub_rows (keep highest-ranked occurrence).
        seen_pub = set()
        new_pub_unique = []
        for r in new_pub_rows:
            p = r["发行商"]
            if p in seen_pub:
                continue
            seen_pub.add(p)
            new_pub_unique.append(r)

        ws_sum.append([sheet_name, len(new_games), len(new_pub_unique)])

        ws = wb.create_sheet(title=sheet_name[:31])
        ws.append(["—— 新出现的游戏 ——"])
        ws.append(headers)
        for r in new_games:
            ws.append([
                r.get("排名"), r.get("游戏名"), r.get("分类"),
                r.get("子类"), r.get("标语"),
                r.get("发行商"), r.get("变化"),
            ])
        ws.append([])
        ws.append(["—— 新出现的发行商（首次进入此榜）——"])
        ws.append(["发行商", "代表游戏", "排名", "分类", "子类"])
        for r in new_pub_unique:
            ws.append([
                r.get("发行商"), r.get("游戏名"), r.get("排名"),
                r.get("分类"), r.get("子类"),
            ])
        # Column widths.
        widths = [8, 24, 12, 16, 32, 28, 14]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # Style summary
    for col, w in enumerate([24, 36], start=1):
        ws_sum.column_dimensions[ws_sum.cell(row=1, column=col).column_letter].width = w
    ws_sum.column_dimensions["C"].width = 14

    delta_str = _human_delta(t_old, t_new)
    out_path = out_dir / (
        f"diff_{t_old.strftime('%Y%m%d_%H%M%S')}_to_"
        f"{t_new.strftime('%Y%m%d_%H%M%S')}__{delta_str}.xlsx"
    )
    wb.save(out_path)
    log(f"[diff] 写出 {out_path}")
    return out_path


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--login", action="store_true",
                    help="Open a browser to log in; save session and exit.")
    ap.add_argument("--anon", action="store_true",
                    help="Force anonymous scrape even if auth file exists.")
    ap.add_argument("--top", type=int, default=None,
                    help="How many rows per board (default: 100 if logged in, "
                         "20 if anonymous).")
    ap.add_argument("--out", type=Path, default=DEFAULT_OUT_DIR,
                    help="Output directory for JSON / XLSX.")
    ap.add_argument("--auth", type=Path, default=DEFAULT_AUTH_FILE,
                    help="Path to login session file.")
    args = ap.parse_args()

    if args.login:
        do_login(auth_file=args.auth)
        return

    data = do_scrape(args.top, args.anon,
                     out_dir=args.out, auth_file=args.auth)
    json_path, xlsx_path = write_outputs(data, out_dir=args.out)
    print(f"Wrote {json_path}")
    print(f"Wrote {xlsx_path}")


if __name__ == "__main__":
    main()
